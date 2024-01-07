import os
import sys
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

API_KEY = '89ded6580277b0ebfac537dd29f143e8'
IMGBB_API_URL = 'https://api.imgbb.com/1/upload'

def upload_to_imgbb(image_path):
    with open(image_path, 'rb') as file:
        files = {'image': (os.path.basename(image_path), file)}
        params = {'key': API_KEY}
        response = requests.post(IMGBB_API_URL, params=params, files=files)

        print(f"Response status code: {response.status_code}")
        print(f"Response content: {response.content}")

        if response.status_code == 200:
            return response.json()['data']['url']
        else:
            return None

def process_images(folder_path, excel_file_path):
    links_dict = {}  # Diccionario para almacenar los enlaces de las imágenes por SKU

    for root, dirs, files in os.walk(folder_path):
        for file_name in files:
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                image_path = os.path.join(root, file_name)
                imgbb_url = upload_to_imgbb(image_path)

                # Extraer el SKU del nombre de la carpeta
                sku = os.path.basename(root)

                if sku not in links_dict:
                    links_dict[sku] = []

                links_dict[sku].append(imgbb_url)

    # Cargar el archivo Excel
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Buscar el código SKU en la columna E y agregar la URL en la columna D
    for sku, image_links in links_dict.items():
        for row in sheet.iter_rows(min_row=2, max_col=5, max_row=sheet.max_row):
            if row[4].value == sku:  # Comparamos con el código SKU en la columna E (index 4)
                # Concatenar los enlaces separados por ", "
                links_str = ', '.join(image_links)
                sheet.cell(row=row[0].row, column=4, value=links_str).alignment = Alignment(wrap_text=True)


    # Guardar el archivo Excel
    workbook.save(excel_file_path)
    print(f"Enlaces guardados en {excel_file_path}")

def obtener_directorio_ejecutable():
    if getattr(sys, 'frozen', False):
        # Estamos en un entorno congelado (como PyInstaller)
        return os.path.dirname(sys.executable)
    else:
        # Estamos en un entorno de script
        return os.path.dirname(os.path.abspath(__file__))


def main():
    # Obtén la ruta del directorio donde se encuentra el .exe
    exe_directory = obtener_directorio_ejecutable()
    #print(exe_directory)
    #input("continuar")

    # Establece image_folder_path como el directorio del .exe
    image_folder_path = exe_directory
    #print(image_folder_path)
    #input("continuar")

    # Construye la ruta completa del archivo Excel en el mismo directorio
    excel_file_path = os.path.join(exe_directory, 'Publicar.xlsx')
    #print(excel_file_path)
    #input("continuar")

    #image_folder_path = 'C:/Users/Usuario/Desktop/img'
    #excel_file_path = 'C:/Users/Usuario/Desktop/img/Publicar.xlsx'

    process_images(image_folder_path, excel_file_path)
    input("Presiona Enter para salir.")



if __name__ == '__main__':
    main()

